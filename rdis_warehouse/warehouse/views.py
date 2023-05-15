from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import Storage, InventoryType, StorageLocation
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.styles.numbers import NumberFormat


def main_view(request):
    stuff = Storage.objects.filter(inventory__type__user=request.user.id, inventory__is_active=True)
    return render(request, 'table.html', {'stuff': stuff})


def export_excel(request):
    if request.method == 'POST':
        # Получаем количество записей, которые нужно экспортировать
        quantity = int(request.POST.get('quantity'))

        # Получаем записи из базы данных
        records = Storage.objects.filter(inventory__type__user=request.user.id,
                                         inventory__is_active=True).order_by('-id_storage')[:quantity]

        # Создаем новый Excel-файл и добавляем в него записи
        wb = Workbook()
        ws = wb.active

        ws.append(["Расположение", "Тип инвентаря",  "Название", "Инвентарный номер",
                   "Характеристики", "Закрепленный сотрудник", "Дата начала", "Дата окончания",
                   "Комментарий"])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30

        # Устанавливаем формат ячейки для заголовков
        header_font = Font(name='Times New Roman', size=14, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        for record in records:
            ws.append([record.location.location_name, record.inventory.type.type_name,
                       record.inventory.inventory_name, record.inventory.inventory_number,
                       record.inventory.specifications,
                       record.worker.first_name + " " + record.worker.name,
                       record.start_date, record.end_date, record.comment])

        #date_style = NamedStyle(name='date_style')
        #date_style.number_format = NumberFormat('dd.mm.yyyy')

        #for row in ws.iter_rows(min_row=2, max_row=len(records)+1, min_col=8, max_col=9):
        #    for cell in row:
        #        cell.style = date_style

        # Возвращаем Excel-файл в ответе на запрос
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        wb.save(response)
        return response


def my_view(request, pk):
    storage = Storage.objects.get(pk=pk)
    return render(request, 'my_template.html', {'object': storage})


def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('warehouse:main')
        else:
            messages.error(request, "Данные не верны, пожалуйста повторите вход")
            return redirect('warehouse:login')
    else:
        return render(request, 'registration/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, "Вы вышли из системы, чтобы воспользоваться сайтом войдите снова")
    return redirect('warehouse:login')
